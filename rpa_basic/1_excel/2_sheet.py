from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # create new sheet with default name
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet") # create with name
ws2 = wb.create_sheet("NewSheet", 2) # create index at 2nd place

new_ws = wb["NewSheet"] # access sheet with dict type

print(wb.sheetnames) # print all sheet in files

new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
