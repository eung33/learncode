from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# insert date by one line
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 컬럼만 가지고 오기
# print(col_B)

# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"] # B~C 컬럼 가져오기

# for cols in col_range:
#     for cell in cols:
#         print(cell.value)
#     print()

# row_title = ws[1]
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] #2~6번째 줄
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()
from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[1:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()

# #전체 row
# print(tuple(ws.rows))

# #전체 columns
# print(tuple(ws.columns))


for row in tuple(ws.rows):
    for cell in row:
        print(cell.value, end=" ")
    print()


wb.save("sample.xlsx")

